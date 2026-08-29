"""网易云音乐API服务主程序

提供网易云音乐相关API服务，包括：
- 歌曲信息获取
- 音乐搜索
- 歌单和专辑详情
- 音乐下载
- 健康检查
"""

import logging
import re
import shutil
import sys
import tempfile
import time
import traceback
import zipfile
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Dict, Any, Iterable, List, Optional, Tuple
from urllib.parse import parse_qs, quote, urljoin, urlparse
from flask import Flask, request, send_file, render_template, Response


VALID_QUALITIES = (
    'standard', 'exhigh', 'lossless', 'hires',
    'sky', 'jyeffect', 'jymaster', 'dolby'
)

try:
    from music_api import (
        NeteaseAPI, APIException, QualityLevel,
        url_v1, name_v1, lyric_v1, search_music,
        playlist_detail, album_detail,
        load_cookies,
    )
    from music_downloader import MusicDownloader, DownloadException, AudioFormat
except ImportError as e:
    print(f"导入模块失败: {e}")
    print("请确保所有依赖模块存在且可用")
    sys.exit(1)


@dataclass
class APIConfig:
    """API配置类"""
    host: str = '0.0.0.0'
    port: int = 5000
    debug: bool = False
    downloads_dir: str = 'downloads'
    max_file_size: int = 500 * 1024 * 1024  # 500MB
    request_timeout: int = 30
    log_level: str = 'INFO'
    cors_origins: str = '*'
    max_upload_size: int = 10 * 1024 * 1024  # 10MB
    max_batch_items: int = 200
    max_file_items: int = 2000
    batch_workers: int = 4


class APIResponse:
    """API响应工具类"""
    
    @staticmethod
    def success(data: Any = None, message: str = 'success', status_code: int = 200) -> Tuple[Dict[str, Any], int]:
        """成功响应"""
        response = {
            'status': status_code,
            'success': True,
            'message': message
        }
        if data is not None:
            response['data'] = data
        return response, status_code
    
    @staticmethod
    def error(message: str, status_code: int = 400, error_code: str = None) -> Tuple[Dict[str, Any], int]:
        """错误响应"""
        response = {
            'status': status_code,
            'success': False,
            'message': message
        }
        if error_code:
            response['error_code'] = error_code
        return response, status_code


class MusicAPIService:
    """音乐API服务类"""
    
    def __init__(self, config: APIConfig):
        self.config = config
        self.logger = self._setup_logger()
        self.netease_api = NeteaseAPI()
        self.downloader = MusicDownloader()

        # 创建下载目录
        self.downloads_path = Path(config.downloads_dir)
        self.downloads_path.mkdir(exist_ok=True)

        self.logger.info(f"音乐API服务初始化完成，下载目录: {self.downloads_path.absolute()}")
    
    def _setup_logger(self) -> logging.Logger:
        """设置日志记录器"""
        logger = logging.getLogger('music_api')
        logger.setLevel(getattr(logging, self.config.log_level.upper()))
        
        if not logger.handlers:
            # 控制台处理器
            console_handler = logging.StreamHandler()
            console_formatter = logging.Formatter(
                '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
            )
            console_handler.setFormatter(console_formatter)
            logger.addHandler(console_handler)
            
            # 文件处理器
            try:
                file_handler = logging.FileHandler('music_api.log', encoding='utf-8')
                file_formatter = logging.Formatter(
                    '%(asctime)s - %(name)s - %(levelname)s - %(funcName)s:%(lineno)d - %(message)s'
                )
                file_handler.setFormatter(file_formatter)
                logger.addHandler(file_handler)
            except Exception as e:
                logger.warning(f"无法创建日志文件: {e}")
        
        return logger
    
    def _get_cookies(self) -> Dict[str, str]:
        """直接从 cookie.txt 读取并解析为 dict"""
        try:
            cookies = load_cookies()
            if not cookies:
                self.logger.warning("cookie.txt 为空或不存在，部分接口（VIP 音质）将不可用")
            return cookies
        except Exception as e:
            self.logger.error(f"读取 cookie.txt 异常: {e}")
            return {}
    
    def _extract_music_id(self, id_or_url: str) -> str:
        """从纯数字或网易云歌曲链接中提取歌曲 ID。"""
        candidate = str(id_or_url or '').strip()
        if re.fullmatch(r'\d+', candidate):
            return candidate

        parsed = urlparse(candidate)
        hostname = (parsed.hostname or '').lower()

        # 仅允许真正的网易云短链域名，避免把包含该字符串的任意 URL 当成短链请求。
        if hostname == '163cn.tv':
            import requests
            response = requests.get(candidate, allow_redirects=False, timeout=10)
            location = response.headers.get('Location')
            if not location:
                raise ValueError("网易云短链接没有返回有效跳转地址")
            candidate = urljoin(candidate, location)
            parsed = urlparse(candidate)
            hostname = (parsed.hostname or '').lower()

        if hostname != 'music.163.com' and not hostname.endswith('.music.163.com'):
            raise ValueError("只支持纯歌曲 ID 或网易云歌曲链接")

        fragment = parsed.fragment or ''
        route_text = f"{parsed.path}?{parsed.query}#{fragment}"
        if '/song' not in route_text:
            raise ValueError("链接不是网易云单曲链接")

        query_parts = [parsed.query]
        if '?' in fragment:
            query_parts.append(fragment.split('?', 1)[1])
        for query_part in query_parts:
            music_ids = parse_qs(query_part).get('id', [])
            if music_ids and re.fullmatch(r'\d+', music_ids[0]):
                return music_ids[0]

        match = re.search(r'(?:[?&#]|^)id=(\d+)', candidate)
        if match:
            return match.group(1)
        raise ValueError("歌曲链接中没有找到有效 ID")

    def extract_music_ids_from_text(self, text: str) -> List[str]:
        """从文本或 CSV 内容中提取歌曲 ID，并按出现顺序去重。"""
        if not text:
            return []

        found: List[str] = []
        seen = set()
        url_pattern = re.compile(r'https?://[^\s<>"\']+', re.IGNORECASE)

        def add_music_id(value: str) -> None:
            if value not in seen:
                seen.add(value)
                found.append(value)

        for line in text.splitlines():
            urls = url_pattern.findall(line)
            for raw_url in urls:
                cleaned_url = raw_url.rstrip('.,，。;；)）]}')
                try:
                    add_music_id(self._extract_music_id(cleaned_url))
                except ValueError:
                    continue

            # URL 已经单独处理，避免再把 URL 参数里的数字提取一次。
            plain_text = url_pattern.sub(' ', line)
            for music_id in re.findall(r'(?<![\w])\d{3,20}(?![\w])', plain_text):
                add_music_id(music_id)

        return found

    def read_song_ids_from_upload(self, uploaded_file) -> List[str]:
        """读取上传的 TXT/CSV/XLSX 文件并提取歌曲 ID。"""
        filename = uploaded_file.filename or ''
        suffix = Path(filename).suffix.lower()
        if suffix not in {'.txt', '.csv', '.xlsx'}:
            raise ValueError("仅支持 .txt、.csv 或 .xlsx 文件")

        raw = uploaded_file.stream.read(self.config.max_upload_size + 1)
        if len(raw) > self.config.max_upload_size:
            raise ValueError("上传文件不能超过 10MB")
        if not raw:
            raise ValueError("上传文件为空")

        if suffix == '.xlsx':
            music_ids = self.extract_music_ids_from_xlsx(raw)
            return self.validate_batch_ids(music_ids)

        text = None
        for encoding in ('utf-8-sig', 'gb18030'):
            try:
                text = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise ValueError("文件编码无法识别，请使用 UTF-8 或 GB18030")

        music_ids = self.extract_music_ids_from_text(text)
        if not music_ids:
            raise ValueError("文件中没有找到歌曲 ID 或网易云单曲链接")
        return self.validate_batch_ids(music_ids)

    def read_song_names_from_upload(self, uploaded_file) -> List[str]:
        """读取上传的 TXT 文件，每行作为一个歌曲搜索词。"""
        filename = uploaded_file.filename or ''
        if Path(filename).suffix.lower() != '.txt':
            raise ValueError("歌名列表仅支持 .txt 文件")

        raw = uploaded_file.stream.read(self.config.max_upload_size + 1)
        if len(raw) > self.config.max_upload_size:
            raise ValueError("上传文件不能超过 10MB")
        if not raw:
            raise ValueError("上传文件为空")

        for encoding in ('utf-8-sig', 'gb18030'):
            try:
                return self.validate_song_queries([raw.decode(encoding)])
            except UnicodeDecodeError:
                continue
        raise ValueError("文件编码无法识别，请使用 UTF-8 或 GB18030")

    def validate_song_queries(self, values: Iterable[Any]) -> List[str]:
        """规范化、去重并限制批量歌名搜索词。"""
        queries: List[str] = []
        seen = set()
        for value in values:
            for line in str(value or '').splitlines():
                query = line.strip().lstrip('\ufeff')
                if not query:
                    continue
                if len(query) > 200:
                    raise ValueError("单个歌曲名称不能超过 200 个字符")
                normalized = query.casefold()
                if normalized not in seen:
                    seen.add(normalized)
                    queries.append(query)

        if not queries:
            raise ValueError("请输入歌曲名称或上传 TXT 歌名列表")
        if len(queries) > self.config.max_batch_items:
            raise ValueError(
                f"一次最多查找 {self.config.max_batch_items} 个歌曲名称，"
                f"当前识别到 {len(queries)} 个"
            )
        return queries

    def resolve_song_name(self, query: str, cookies: Dict[str, str]) -> Dict[str, Any]:
        """查找一个歌曲名称并返回最相关的一首，单条失败不影响其余结果。"""
        try:
            search_query = re.sub(r'\s*[|｜]\s*', ' ', query).strip()
            songs = search_music(search_query, cookies, 10)
            if not songs:
                return {
                    'query': query,
                    'success': False,
                    'error': '未找到相关歌曲',
                }

            ranked_songs = sorted(
                enumerate(songs),
                key=lambda pair: (self._score_song_match(query, pair[1]), -pair[0]),
                reverse=True,
            )
            candidates: List[Dict[str, Any]] = []
            seen_ids = set()
            explicit_parts = re.split(r'[|｜]', query, maxsplit=1)
            expected_artist = re.sub(
                r'[^\w\u4e00-\u9fff]+',
                '',
                explicit_parts[1].casefold(),
            ) if len(explicit_parts) > 1 else ''
            for _, song in ranked_songs:
                music_id = str(song.get('id', '')).strip()
                if not music_id.isdigit() or music_id in seen_ids:
                    continue
                seen_ids.add(music_id)
                candidate_artists = str(song.get('artists', ''))
                normalized_artists = re.sub(
                    r'[^\w\u4e00-\u9fff]+', '', candidate_artists.casefold()
                )
                candidates.append({
                    'id': music_id,
                    'name': str(song.get('name', '')),
                    'artists': candidate_artists,
                    'album': str(song.get('album', '')),
                    'picUrl': str(song.get('picUrl', '')),
                    'link': f"https://music.163.com/song?id={music_id}",
                    'artist_matched': (
                        expected_artist in normalized_artists
                        if expected_artist
                        else None
                    ),
                })
                if len(candidates) >= 5:
                    break
            if not candidates:
                raise ValueError("搜索结果缺少有效歌曲 ID")

            song = candidates[0]
            return {
                'query': query,
                'success': True,
                **song,
                'candidate_count': len(candidates),
                'candidates': candidates,
                'auto_selected': song.get('artist_matched') is not False,
            }
        except Exception as e:
            return {
                'query': query,
                'success': False,
                'error': str(e),
            }

    @staticmethod
    def _score_song_match(query: str, song: Dict[str, Any]) -> int:
        """按歌名和歌手覆盖度给搜索候选排序，降低翻唱/伴奏误匹配。"""
        normalize = lambda value: re.sub(
            r'[^\w\u4e00-\u9fff]+', '', str(value or '').casefold()
        )
        explicit_parts = re.split(r'[|｜]', query, maxsplit=1)
        explicit_title = normalize(explicit_parts[0]) if explicit_parts else ''
        explicit_artist = normalize(explicit_parts[1]) if len(explicit_parts) > 1 else ''
        query_tokens = [
            normalize(token)
            for token in re.split(r'[\s|｜]+', query)
            if token
        ]
        query_tokens = [token for token in query_tokens if token]
        name = normalize(song.get('name'))
        artists = normalize(song.get('artists'))
        album = normalize(song.get('album'))
        searchable = f"{name}{artists}{album}"
        if not query_tokens:
            return 0

        score = 0
        title_token = explicit_title or query_tokens[0]
        if name == title_token:
            score += 40
        elif name.startswith(title_token):
            score += 20

        missing_tokens = 0
        for token in query_tokens:
            if token == name:
                score += 15
            elif token in name:
                score += 8
            if token == artists:
                score += 40
            elif token in artists:
                score += 16
            if token in searchable:
                score += 10
            else:
                missing_tokens += 1
        score -= missing_tokens * 30

        if explicit_artist:
            if artists == explicit_artist:
                score += 100
            elif explicit_artist in artists:
                score += 60
            else:
                score -= 80

        query_text = normalize(query)
        for marker in ('翻唱', 'cover', '伴奏', '纯音乐', '钢琴版', '吉他版', 'dj版', '原唱'):
            normalized_marker = normalize(marker)
            if normalized_marker in name and normalized_marker not in query_text:
                score -= 15
        return score

    def extract_music_ids_from_xlsx(self, raw: bytes) -> List[str]:
        """从带“歌曲 ID”或“网易云链接”表头的 Excel 工作表中提取歌曲。"""
        try:
            with zipfile.ZipFile(BytesIO(raw)) as archive:
                expanded_size = sum(item.file_size for item in archive.infolist())
                if expanded_size > 100 * 1024 * 1024:
                    raise ValueError("Excel 解压后体积过大，请精简文件后重试")
        except zipfile.BadZipFile as e:
            raise ValueError("Excel 文件已损坏或不是有效的 .xlsx 文件") from e

        try:
            from openpyxl import load_workbook
            workbook = load_workbook(
                BytesIO(raw),
                read_only=False,
                data_only=True,
                keep_links=False,
            )
        except Exception as e:
            raise ValueError(f"无法读取 Excel 文件: {e}") from e

        found: List[str] = []
        seen = set()
        matched_sheet = False

        def add_candidate(value: Any) -> bool:
            if value is None or isinstance(value, bool):
                return False
            if isinstance(value, float) and value.is_integer():
                value = int(value)
            candidate = str(value).strip()
            if not candidate:
                return False
            try:
                music_id = self._extract_music_id(candidate)
            except ValueError:
                return False
            if music_id not in seen:
                seen.add(music_id)
                found.append(music_id)
            return True

        try:
            for sheet in workbook.worksheets:
                header_row = None
                id_columns: List[int] = []
                link_columns: List[int] = []
                scan_rows = min(sheet.max_row or 0, 100)
                scan_columns = min(sheet.max_column or 0, 200)

                for row in sheet.iter_rows(
                    min_row=1,
                    max_row=scan_rows,
                    max_col=scan_columns,
                ):
                    current_id_columns: List[int] = []
                    current_link_columns: List[int] = []
                    for cell in row:
                        header = re.sub(r'\s+', '', str(cell.value or '')).lower()
                        if header in {'歌曲id', '音乐id', '网易云歌曲id'}:
                            current_id_columns.append(cell.column)
                        elif header in {'网易云链接', '歌曲链接', '音乐链接'}:
                            current_link_columns.append(cell.column)
                    if current_id_columns or current_link_columns:
                        header_row = row[0].row
                        id_columns = current_id_columns
                        link_columns = current_link_columns
                        break

                if header_row is None:
                    continue
                matched_sheet = True
                max_column = max(id_columns + link_columns)

                for row in sheet.iter_rows(
                    min_row=header_row + 1,
                    max_row=sheet.max_row,
                    max_col=max_column,
                ):
                    added = False
                    for column in link_columns:
                        cell = row[column - 1]
                        hyperlink = getattr(cell.hyperlink, 'target', None) if cell.hyperlink else None
                        for value in (hyperlink, cell.value):
                            if add_candidate(value):
                                added = True
                                break
                        if added:
                            break
                    if added:
                        continue
                    for column in id_columns:
                        if add_candidate(row[column - 1].value):
                            break
        finally:
            workbook.close()

        if not matched_sheet:
            raise ValueError("Excel 中没有找到“歌曲 ID”或“网易云链接”表头")
        if not found:
            raise ValueError("Excel 中没有找到有效的歌曲 ID 或网易云单曲链接")
        return found

    def validate_batch_ids(self, values: Iterable[Any]) -> List[str]:
        """规范化、去重并限制批量歌曲 ID。"""
        music_ids: List[str] = []
        seen = set()
        for value in values:
            value_text = str(value or '').strip()
            if not value_text:
                continue
            extracted = self.extract_music_ids_from_text(value_text)
            if not extracted:
                try:
                    extracted = [self._extract_music_id(value_text)]
                except ValueError:
                    continue
            for music_id in extracted:
                if music_id not in seen:
                    seen.add(music_id)
                    music_ids.append(music_id)

        if not music_ids:
            raise ValueError("没有提供有效的歌曲 ID")
        if len(music_ids) > self.config.max_file_items:
            raise ValueError(
                f"单个文件或请求最多识别 {self.config.max_file_items} 首歌曲，"
                f"当前识别到 {len(music_ids)} 首"
            )
        return music_ids

    def split_batch_ids(self, music_ids: List[str]) -> List[List[str]]:
        """按配置的单批上限切分歌曲 ID。"""
        return [
            music_ids[index:index + self.config.max_batch_items]
            for index in range(0, len(music_ids), self.config.max_batch_items)
        ]

    def parse_song_summary(self, music_id: str, level: str, cookies: Dict[str, str]) -> Dict[str, Any]:
        """解析批量列表需要的歌曲摘要，单首失败不会中断整个批次。"""
        try:
            song_info = name_v1(int(music_id))
            url_info = url_v1(int(music_id), level, cookies)
            songs = song_info.get('songs', []) if song_info else []
            if not songs:
                raise APIException("未找到歌曲信息")

            song_data = songs[0]
            url_data = (url_info.get('data') or [{}])[0] if url_info else {}
            if not url_data.get('url'):
                raise APIException("未获取到下载链接，请检查 Cookie、歌曲版权或所选音质")
            return {
                'id': music_id,
                'success': True,
                'name': song_data.get('name', ''),
                'artists': ', '.join(artist.get('name', '') for artist in song_data.get('ar', [])),
                'album': song_data.get('al', {}).get('name', ''),
                'pic': song_data.get('al', {}).get('picUrl', ''),
                'url': url_data.get('url', ''),
                'level': url_data.get('level', level),
                'size': url_data.get('size', 0),
                'size_formatted': self._format_file_size(url_data.get('size', 0)),
                'file_type': url_data.get('type', ''),
            }
        except Exception as e:
            return {
                'id': music_id,
                'success': False,
                'error': str(e),
            }
    
    def _format_file_size(self, size_bytes: int) -> str:
        """格式化文件大小"""
        if size_bytes == 0:
            return "0B"
        
        units = ["B", "KB", "MB", "GB", "TB"]
        size = float(size_bytes)
        unit_index = 0
        
        while size >= 1024.0 and unit_index < len(units) - 1:
            size /= 1024.0
            unit_index += 1
        
        return f"{size:.2f}{units[unit_index]}"
    
    def _get_quality_display_name(self, quality: str) -> str:
        """获取音质显示名称"""
        quality_names = {
            'standard': "标准音质",
            'exhigh': "极高音质", 
            'lossless': "无损音质",
            'hires': "Hi-Res音质",
            'sky': "沉浸环绕声",
            'jyeffect': "高清环绕声",
            'jymaster': "超清母带",
            'dolby': "杜比全景声"
        }
        return quality_names.get(quality, f"未知音质({quality})")
    
    def _validate_request_params(self, required_params: Dict[str, Any]) -> Optional[Tuple[Dict[str, Any], int]]:
        """验证请求参数"""
        for param_name, param_value in required_params.items():
            if not param_value:
                return APIResponse.error(f"参数 '{param_name}' 不能为空", 400)
        return None
    
    def _safe_get_request_data(self) -> Dict[str, Any]:
        """安全获取请求数据"""
        try:
            if request.method == 'GET':
                return dict(request.args)
            else:
                # 优先使用JSON数据，然后是表单数据
                json_data = request.get_json(silent=True) or {}
                form_data = dict(request.form)
                # 合并数据，JSON优先
                return {**form_data, **json_data}
        except Exception as e:
            self.logger.error(f"获取请求数据失败: {e}")
            return {}


# 创建Flask应用和服务实例
config = APIConfig()
app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = config.max_upload_size + 64 * 1024
api_service = MusicAPIService(config)


@app.before_request
def before_request():
    """请求前处理"""
    # 记录请求信息
    api_service.logger.info(
        f"{request.method} {request.path} - IP: {request.remote_addr} - "
        f"User-Agent: {request.headers.get('User-Agent', 'Unknown')}"
    )


@app.after_request
def after_request(response: Response) -> Response:
    """请求后处理 - 设置CORS头"""
    response.headers.add('Access-Control-Allow-Origin', config.cors_origins)
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,POST,OPTIONS')
    response.headers.add('Access-Control-Max-Age', '3600')
    response.headers.add(
        'Access-Control-Expose-Headers',
        'Content-Disposition,X-Download-Filename,X-Batch-Succeeded,X-Batch-Failed'
    )
    
    # 记录响应信息
    api_service.logger.info(f"响应状态: {response.status_code}")
    return response


@app.errorhandler(400)
def handle_bad_request(e):
    """处理400错误"""
    return APIResponse.error("请求参数错误", 400)


@app.errorhandler(404)
def handle_not_found(e):
    """处理404错误"""
    return APIResponse.error("请求的资源不存在", 404)


@app.errorhandler(500)
def handle_internal_error(e):
    """处理500错误"""
    api_service.logger.error(f"服务器内部错误: {e}")
    return APIResponse.error("服务器内部错误", 500)


@app.errorhandler(413)
def handle_file_too_large(e):
    """处理上传体积超限。"""
    return APIResponse.error("上传文件不能超过 10MB", 413)


@app.route('/')
def index() -> str:
    """首页路由"""
    return render_template('index.html')


@app.route('/health', methods=['GET'])
def health_check():
    """健康检查API"""
    try:
        # 检查 Cookie 状态：直接读 cookie.txt
        cookies = load_cookies()
        cookie_status = 'valid' if cookies else 'invalid'

        health_info = {
            'service': 'running',
            'timestamp': int(time.time()) if 'time' in sys.modules else None,
            'cookie_status': cookie_status,
            'cookie_count': len(cookies),
            'downloads_dir': str(api_service.downloads_path.absolute()),
            'version': '2.1.0'
        }

        return APIResponse.success(health_info, "API服务运行正常")
        
    except Exception as e:
        api_service.logger.error(f"健康检查失败: {e}")
        return APIResponse.error(f"健康检查失败: {str(e)}", 500)


@app.route('/song', methods=['GET', 'POST'])
@app.route('/Song_V1', methods=['GET', 'POST'])  # 向后兼容
def get_song_info():
    """获取歌曲信息API"""
    try:
        # 获取请求参数
        data = api_service._safe_get_request_data()
        song_ids = data.get('ids') or data.get('id')
        url = data.get('url')
        level = data.get('level', 'lossless')
        info_type = data.get('type', 'url')
        
        # 参数验证
        if not song_ids and not url:
            return APIResponse.error("必须提供 'ids'、'id' 或 'url' 参数")
        
        # 提取音乐ID
        music_id = api_service._extract_music_id(song_ids or url)
        
        # 验证音质参数
        if level not in VALID_QUALITIES:
            return APIResponse.error(f"无效的音质参数，支持: {', '.join(VALID_QUALITIES)}")
        
        # 验证类型参数
        valid_types = ['url', 'name', 'lyric', 'json']
        if info_type not in valid_types:
            return APIResponse.error(f"无效的类型参数，支持: {', '.join(valid_types)}")
        
        cookies = api_service._get_cookies()
        
        # 根据类型获取不同信息
        if info_type == 'url':
            result = url_v1(music_id, level, cookies)
            if result and result.get('data') and len(result['data']) > 0:
                song_data = result['data'][0]
                response_data = {
                    'id': song_data.get('id'),
                    'url': song_data.get('url'),
                    'level': song_data.get('level'),
                    'quality_name': api_service._get_quality_display_name(song_data.get('level', level)),
                    'size': song_data.get('size'),
                    'size_formatted': api_service._format_file_size(song_data.get('size', 0)),
                    'type': song_data.get('type'),
                    'bitrate': song_data.get('br')
                }
                return APIResponse.success(response_data, "获取歌曲URL成功")
            else:
                return APIResponse.error("获取音乐URL失败，可能是版权限制或音质不支持", 404)
        
        elif info_type == 'name':
            result = name_v1(music_id)
            return APIResponse.success(result, "获取歌曲信息成功")
        
        elif info_type == 'lyric':
            result = lyric_v1(music_id, cookies)
            return APIResponse.success(result, "获取歌词成功")
        
        elif info_type == 'json':
            # 获取完整的歌曲信息（用于前端解析）
            song_info = name_v1(music_id)
            url_info = url_v1(music_id, level, cookies)
            lyric_info = lyric_v1(music_id, cookies)
            
            if not song_info or 'songs' not in song_info or not song_info['songs']:
                return APIResponse.error("未找到歌曲信息", 404)
            
            song_data = song_info['songs'][0]
            
            # 构建前端期望的响应格式
            response_data = {
                'id': music_id,
                'name': song_data.get('name', ''),
                'ar_name': ', '.join(artist['name'] for artist in song_data.get('ar', [])),
                'al_name': song_data.get('al', {}).get('name', ''),
                'pic': song_data.get('al', {}).get('picUrl', ''),
                'level': level,
                'lyric': lyric_info.get('lrc', {}).get('lyric', '') if lyric_info else '',
                'tlyric': lyric_info.get('tlyric', {}).get('lyric', '') if lyric_info else ''
            }
            
            # 添加URL和大小信息
            if url_info and url_info.get('data') and len(url_info['data']) > 0:
                url_data = url_info['data'][0]
                response_data.update({
                    'url': url_data.get('url', ''),
                    'size': api_service._format_file_size(url_data.get('size', 0)),
                    'level': url_data.get('level', level)
                })
            else:
                response_data.update({
                    'url': '',
                    'size': '获取失败'
                })
            
            return APIResponse.success(response_data, "获取歌曲信息成功")
            
    except ValueError as e:
        return APIResponse.error(str(e), 400)
    except APIException as e:
        api_service.logger.error(f"API调用失败: {e}")
        return APIResponse.error(f"API调用失败: {str(e)}", 500)
    except Exception as e:
        api_service.logger.error(f"获取歌曲信息异常: {e}\n{traceback.format_exc()}")
        return APIResponse.error(f"服务器错误: {str(e)}", 500)


def _get_batch_ids_from_request() -> List[str]:
    """从上传文件、JSON 或表单字段读取批量歌曲 ID。"""
    uploaded_file = request.files.get('file')
    if uploaded_file and uploaded_file.filename:
        return api_service.read_song_ids_from_upload(uploaded_file)

    data = api_service._safe_get_request_data()
    raw_ids = data.get('ids') or data.get('content')
    if isinstance(raw_ids, list):
        values = raw_ids
    elif raw_ids is not None:
        values = [raw_ids]
    else:
        values = []
    return api_service.validate_batch_ids(values)


def _get_song_queries_from_request() -> List[str]:
    """从 TXT 上传和文本输入中读取批量歌曲名称。"""
    values: List[str] = []
    uploaded_file = request.files.get('file')
    if uploaded_file and uploaded_file.filename:
        values.extend(api_service.read_song_names_from_upload(uploaded_file))

    data = api_service._safe_get_request_data()
    raw_queries = data.get('queries') or data.get('content')
    if isinstance(raw_queries, list):
        values.extend(str(value) for value in raw_queries)
    elif raw_queries is not None:
        values.append(str(raw_queries))
    return api_service.validate_song_queries(values)


@app.route('/batch/resolve-names', methods=['POST'])
def batch_resolve_song_names():
    """批量搜索歌曲名称，返回最相关歌曲的 ID 和网易云链接。"""
    try:
        queries = _get_song_queries_from_request()
        cookies = api_service._get_cookies()

        def resolve_one(query: str) -> Dict[str, Any]:
            return api_service.resolve_song_name(query, cookies)

        with ThreadPoolExecutor(max_workers=config.batch_workers) as executor:
            results = list(executor.map(resolve_one, queries))

        matched_ids: List[str] = []
        seen_ids = set()
        for item in results:
            music_id = (
                item.get('id')
                if item.get('success') and item.get('auto_selected', True)
                else None
            )
            if music_id and music_id not in seen_ids:
                seen_ids.add(music_id)
                matched_ids.append(music_id)

        succeeded = sum(1 for item in results if item.get('success'))
        needs_review = sum(
            1
            for item in results
            if item.get('success') and not item.get('auto_selected', True)
        )
        response_data = {
            'queries': queries,
            'ids': matched_ids,
            'total': len(queries),
            'succeeded': succeeded,
            'failed': len(queries) - succeeded,
            'needs_review': needs_review,
            'unique_matches': len(matched_ids),
            'results': results,
        }
        return APIResponse.success(
            response_data,
            f"歌名查找完成，找到 {succeeded} 个，其中需确认 {needs_review} 个",
        )
    except ValueError as e:
        return APIResponse.error(str(e), 400)
    except Exception as e:
        api_service.logger.error(f"批量歌名查找异常: {e}\n{traceback.format_exc()}")
        return APIResponse.error(f"批量歌名查找失败: {str(e)}", 500)


@app.route('/batch/parse', methods=['POST'])
def batch_parse_songs():
    """从上传文件或 ID 列表中批量解析歌曲。"""
    try:
        data = api_service._safe_get_request_data()
        level = data.get('level', 'lossless')
        if level not in VALID_QUALITIES:
            return APIResponse.error(f"无效的音质参数，支持: {', '.join(VALID_QUALITIES)}")

        music_ids = _get_batch_ids_from_request()
        cookies = api_service._get_cookies()

        def parse_one(music_id: str) -> Dict[str, Any]:
            return api_service.parse_song_summary(music_id, level, cookies)

        batches = api_service.split_batch_ids(music_ids)
        results: List[Dict[str, Any]] = []
        with ThreadPoolExecutor(max_workers=config.batch_workers) as executor:
            for batch in batches:
                results.extend(executor.map(parse_one, batch))

        succeeded = sum(1 for item in results if item.get('success'))
        response_data = {
            'ids': music_ids,
            'total': len(music_ids),
            'succeeded': succeeded,
            'failed': len(music_ids) - succeeded,
            'level': level,
            'batch_size': config.max_batch_items,
            'batch_count': len(batches),
            'results': results,
        }
        return APIResponse.success(response_data, f"批量解析完成，成功 {succeeded} 首")
    except ValueError as e:
        return APIResponse.error(str(e), 400)
    except Exception as e:
        api_service.logger.error(f"批量解析异常: {e}\n{traceback.format_exc()}")
        return APIResponse.error(f"批量解析失败: {str(e)}", 500)


@app.route('/batch/download', methods=['POST'])
def batch_download_songs():
    """批量下载歌曲并打包成 ZIP。"""
    temp_root: Optional[Path] = None
    try:
        data = api_service._safe_get_request_data()
        quality = data.get('quality', data.get('level', 'lossless'))
        if quality not in VALID_QUALITIES:
            return APIResponse.error(f"无效的音质参数，支持: {', '.join(VALID_QUALITIES)}")

        music_ids = _get_batch_ids_from_request()
        temp_root = Path(tempfile.mkdtemp(prefix='netease_batch_'))

        def download_one(music_id: str) -> Dict[str, Any]:
            song_dir = temp_root / music_id
            try:
                downloader = MusicDownloader(
                    download_dir=str(song_dir),
                    max_file_size=config.max_file_size,
                )
                result = downloader.download_music_file(int(music_id), quality)
                if not result.success or not result.file_path:
                    return {
                        'id': music_id,
                        'success': False,
                        'error': result.error_message or '下载失败',
                    }
                return {
                    'id': music_id,
                    'success': True,
                    'file_path': result.file_path,
                    'filename': Path(result.file_path).name,
                }
            except Exception as e:
                return {'id': music_id, 'success': False, 'error': str(e)}

        batches = api_service.split_batch_ids(music_ids)
        results: List[Dict[str, Any]] = []
        with ThreadPoolExecutor(max_workers=config.batch_workers) as executor:
            for batch in batches:
                results.extend(executor.map(download_one, batch))

        successful = [item for item in results if item.get('success')]
        failed = [item for item in results if not item.get('success')]
        if not successful:
            shutil.rmtree(temp_root, ignore_errors=True)
            temp_root = None
            error_preview = '；'.join(
                f"{item['id']}: {item.get('error', '下载失败')}" for item in failed[:3]
            )
            return APIResponse.error(f"批量下载全部失败：{error_preview}", 502)

        zip_name = f"netease_batch_{time.strftime('%Y%m%d_%H%M%S')}_{quality}.zip"
        zip_path = temp_root / zip_name
        used_names = set()
        with zipfile.ZipFile(zip_path, 'w', compression=zipfile.ZIP_STORED) as archive:
            for item in successful:
                file_path = Path(item['file_path'])
                archive_name = item['filename']
                if archive_name in used_names:
                    archive_name = f"{item['id']} - {archive_name}"
                used_names.add(archive_name)
                archive.write(file_path, arcname=archive_name)

            summary_lines = [
                f"音质：{quality}",
                f"请求：{len(music_ids)} 首",
                f"批次：{len(batches)} 批（每批最多 {config.max_batch_items} 首）",
                f"成功：{len(successful)} 首",
                f"失败：{len(failed)} 首",
                '',
            ]
            if failed:
                summary_lines.append('失败明细：')
                summary_lines.extend(
                    f"- {item['id']}: {item.get('error', '下载失败')}" for item in failed
                )
            archive.writestr('批量下载结果.txt', '\ufeff' + '\n'.join(summary_lines))

        # ZIP 生成后立即清理音频临时目录，只保留待发送的压缩包。
        for child in temp_root.iterdir():
            if child == zip_path:
                continue
            if child.is_dir():
                shutil.rmtree(child, ignore_errors=True)
            else:
                child.unlink(missing_ok=True)

        response = send_file(
            str(zip_path),
            as_attachment=True,
            download_name=zip_name,
            mimetype='application/zip',
        )
        response.headers['X-Batch-Succeeded'] = str(len(successful))
        response.headers['X-Batch-Failed'] = str(len(failed))
        cleanup_root = temp_root
        response.call_on_close(lambda: shutil.rmtree(cleanup_root, ignore_errors=True))
        temp_root = None
        return response
    except ValueError as e:
        if temp_root:
            shutil.rmtree(temp_root, ignore_errors=True)
        return APIResponse.error(str(e), 400)
    except Exception as e:
        if temp_root:
            shutil.rmtree(temp_root, ignore_errors=True)
        api_service.logger.error(f"批量下载异常: {e}\n{traceback.format_exc()}")
        return APIResponse.error(f"批量下载失败: {str(e)}", 500)


@app.route('/search', methods=['GET', 'POST'])
@app.route('/Search', methods=['GET', 'POST'])  # 向后兼容
def search_music_api():
    """搜索音乐API"""
    try:
        # 获取请求参数
        data = api_service._safe_get_request_data()
        keyword = data.get('keyword') or data.get('keywords') or data.get('q')
        limit = int(data.get('limit', 30))
        offset = int(data.get('offset', 0))
        search_type = data.get('type', '1')  # 1-歌曲, 10-专辑, 100-歌手, 1000-歌单
        
        # 参数验证
        validation_error = api_service._validate_request_params({'keyword': keyword})
        if validation_error:
            return validation_error
        
        # 限制搜索数量
        if limit > 100:
            limit = 100
        
        cookies = api_service._get_cookies()
        result = search_music(keyword, cookies, limit)
        
        # search_music返回的是歌曲列表，需要包装成前端期望的格式
        if result:
            for song in result:
                # 添加艺术家字符串（如果需要）
                if 'artists' in song:
                    song['artist_string'] = song['artists']
        
        return APIResponse.success(result, "搜索完成")
        
    except ValueError as e:
        return APIResponse.error(f"参数格式错误: {str(e)}")
    except Exception as e:
        api_service.logger.error(f"搜索音乐异常: {e}\n{traceback.format_exc()}")
        return APIResponse.error(f"搜索失败: {str(e)}", 500)


@app.route('/playlist', methods=['GET', 'POST'])
@app.route('/Playlist', methods=['GET', 'POST'])  # 向后兼容
def get_playlist():
    """获取歌单详情API"""
    try:
        # 获取请求参数
        data = api_service._safe_get_request_data()
        playlist_id = data.get('id')
        
        # 参数验证
        validation_error = api_service._validate_request_params({'playlist_id': playlist_id})
        if validation_error:
            return validation_error
        
        cookies = api_service._get_cookies()
        result = playlist_detail(playlist_id, cookies)
        
        # 适配前端期望的响应格式
        response_data = {
            'status': 'success',
            'playlist': result
        }
        
        return APIResponse.success(response_data, "获取歌单详情成功")
        
    except Exception as e:
        api_service.logger.error(f"获取歌单异常: {e}\n{traceback.format_exc()}")
        return APIResponse.error(f"获取歌单失败: {str(e)}", 500)


@app.route('/album', methods=['GET', 'POST'])
@app.route('/Album', methods=['GET', 'POST'])  # 向后兼容
def get_album():
    """获取专辑详情API"""
    try:
        # 获取请求参数
        data = api_service._safe_get_request_data()
        album_id = data.get('id')
        
        # 参数验证
        validation_error = api_service._validate_request_params({'album_id': album_id})
        if validation_error:
            return validation_error
        
        cookies = api_service._get_cookies()
        result = album_detail(album_id, cookies)
        
        # 适配前端期望的响应格式
        response_data = {
            'status': 200,
            'album': result
        }
        
        return APIResponse.success(response_data, "获取专辑详情成功")
        
    except Exception as e:
        api_service.logger.error(f"获取专辑异常: {e}\n{traceback.format_exc()}")
        return APIResponse.error(f"获取专辑失败: {str(e)}", 500)


@app.route('/download', methods=['GET', 'POST'])
@app.route('/Download', methods=['GET', 'POST'])  # 向后兼容
def download_music_api():
    """下载音乐API"""
    try:
        # 获取请求参数
        data = api_service._safe_get_request_data()
        music_id = data.get('id')
        quality = data.get('quality', 'lossless')
        return_format = data.get('format', 'file')  # file 或 json
        
        # 参数验证
        validation_error = api_service._validate_request_params({'music_id': music_id})
        if validation_error:
            return validation_error
        
        # 验证音质参数
        if quality not in VALID_QUALITIES:
            return APIResponse.error(f"无效的音质参数，支持: {', '.join(VALID_QUALITIES)}")
        
        # 验证返回格式
        if return_format not in ['file', 'json']:
            return APIResponse.error("返回格式只支持 'file' 或 'json'")
        
        music_id = api_service._extract_music_id(music_id)
        cookies = api_service._get_cookies()
        
        # 获取音乐基本信息
        song_info = name_v1(music_id)
        if not song_info or 'songs' not in song_info or not song_info['songs']:
            return APIResponse.error("未找到音乐信息", 404)
        
        # 获取音乐下载链接
        url_info = url_v1(music_id, quality, cookies)
        if not url_info or 'data' not in url_info or not url_info['data'] or not url_info['data'][0].get('url'):
            return APIResponse.error("无法获取音乐下载链接，可能是版权限制或音质不支持", 404)
        
        # 构建音乐信息
        song_data = song_info['songs'][0]
        url_data = url_info['data'][0]
        if url_data.get('size', 0) > config.max_file_size:
            return APIResponse.error(
                f"歌曲文件超过单文件限制 {api_service._format_file_size(config.max_file_size)}",
                413,
            )
        
        music_info = {
            'id': music_id,
            'name': song_data['name'],
            'artist_string': ', '.join(artist['name'] for artist in song_data['ar']),
            'album': song_data['al']['name'],
            'pic_url': song_data['al']['picUrl'],
            'file_type': url_data['type'],
            'file_size': url_data['size'],
            'duration': song_data.get('dt', 0),
            'download_url': url_data['url']
        }
        
        # 下载器负责统一生成包含 ID 和音质的文件名，避免不同音质误用同一缓存文件。
        try:
            download_result = api_service.downloader.download_music_file(music_id, quality)
            if not download_result.success or not download_result.file_path:
                return APIResponse.error(f"下载失败: {download_result.error_message}", 500)
            file_path = Path(download_result.file_path)
            filename = file_path.name
            api_service.logger.info(f"下载完成: {filename}")
        except DownloadException as e:
            api_service.logger.error(f"下载异常: {e}")
            return APIResponse.error(f"下载失败: {str(e)}", 500)
        
        # 根据返回格式返回结果
        if return_format == 'json':
            response_data = {
                'music_id': music_id,
                'name': music_info['name'],
                'artist': music_info['artist_string'],
                'album': music_info['album'],
                'quality': quality,
                'quality_name': api_service._get_quality_display_name(quality),
                'file_type': file_path.suffix.lstrip('.') or music_info['file_type'],
                'file_size': music_info['file_size'],
                'file_size_formatted': api_service._format_file_size(music_info['file_size']),
                'file_path': str(file_path.absolute()),
                'filename': filename,
                'duration': music_info['duration']
            }
            return APIResponse.success(response_data, "下载完成")
        else:
            # 返回文件下载
            if not file_path.exists():
                return APIResponse.error("文件不存在", 404)
            
            try:
                response = send_file(
                    str(file_path),
                    as_attachment=True,
                    download_name=filename,
                    mimetype=f"audio/{file_path.suffix.lstrip('.') or music_info['file_type']}"
                )
                response.headers['X-Download-Message'] = 'Download completed successfully'
                response.headers['X-Download-Filename'] = quote(filename, safe='')
                return response
            except Exception as e:
                api_service.logger.error(f"发送文件失败: {e}")
                return APIResponse.error(f"文件发送失败: {str(e)}", 500)
            
    except ValueError as e:
        return APIResponse.error(str(e), 400)
    except Exception as e:
        api_service.logger.error(f"下载音乐异常: {e}\n{traceback.format_exc()}")
        return APIResponse.error(f"下载异常: {str(e)}", 500)


@app.route('/api/info', methods=['GET'])
def api_info():
    """API信息接口"""
    try:
        info = {
            'name': '网易云音乐API服务',
            'version': '2.1.0',
            'description': '提供网易云音乐相关API服务',
            'endpoints': {
                '/health': 'GET - 健康检查',
                '/song': 'GET/POST - 获取歌曲信息',
                '/search': 'GET/POST - 搜索音乐',
                '/playlist': 'GET/POST - 获取歌单详情',
                '/album': 'GET/POST - 获取专辑详情',
                '/download': 'GET/POST - 下载音乐',
                '/batch/resolve-names': 'POST - 批量歌名查找歌曲 ID 和链接',
                '/batch/parse': 'POST - 上传文件或 ID 列表批量解析',
                '/batch/download': 'POST - 批量下载并打包 ZIP',
                '/api/info': 'GET - API信息'
            },
            'supported_qualities': [
                'standard', 'exhigh', 'lossless', 
                'hires', 'sky', 'jyeffect', 'jymaster', 'dolby'
            ],
            'config': {
                'downloads_dir': str(api_service.downloads_path.absolute()),
                'max_file_size': f"{config.max_file_size // (1024*1024)}MB",
                'max_upload_size': f"{config.max_upload_size // (1024*1024)}MB",
                'max_batch_items': config.max_batch_items,
                'max_file_items': config.max_file_items,
                'request_timeout': f"{config.request_timeout}s"
            }
        }
        
        return APIResponse.success(info, "API信息获取成功")
        
    except Exception as e:
        api_service.logger.error(f"获取API信息异常: {e}")
        return APIResponse.error(f"获取API信息失败: {str(e)}", 500)


def start_api_server():
    """启动API服务器"""
    try:
        print("\n" + "="*60)
        print("🚀 网易云音乐API服务启动中...")
        print("="*60)
        print(f"📡 服务地址: http://{config.host}:{config.port}")
        print(f"📁 下载目录: {api_service.downloads_path.absolute()}")
        print(f"📋 日志级别: {config.log_level}")
        print("\n📚 API端点:")
        print(f"  ├─ GET  /health        - 健康检查")
        print(f"  ├─ POST /song          - 获取歌曲信息")
        print(f"  ├─ POST /search        - 搜索音乐")
        print(f"  ├─ POST /playlist      - 获取歌单详情")
        print(f"  ├─ POST /album         - 获取专辑详情")
        print(f"  ├─ POST /download      - 下载音乐")
        print(f"  ├─ POST /batch/resolve-names - 批量歌名查找")
        print(f"  ├─ POST /batch/parse   - 文件批量解析")
        print(f"  ├─ POST /batch/download - 批量下载 ZIP")
        print(f"  └─ GET  /api/info      - API信息")
        print("\n🎵 支持的音质:")
        print(f"  standard, exhigh, lossless, hires, sky, jyeffect, jymaster, dolby")
        print("="*60)
        print(f"⏰ 启动时间: {time.strftime('%Y-%m-%d %H:%M:%S')}")
        print("🌟 服务已就绪，等待请求...\n")
        
        # 启动Flask应用
        app.run(
            host=config.host,
            port=config.port,
            debug=config.debug,
            threaded=True
        )
        
    except KeyboardInterrupt:
        print("\n\n👋 服务已停止")
    except Exception as e:
        api_service.logger.error(f"启动服务失败: {e}")
        print(f"❌ 启动失败: {e}")
        sys.exit(1)


if __name__ == '__main__':
    start_api_server()

